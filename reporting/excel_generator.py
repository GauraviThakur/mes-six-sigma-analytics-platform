import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_excel_report(oee_df: pd.DataFrame, spc_metrics: dict, output_path: str = "MES_Manufacturing_Performance_Report.xlsx"):
    """Generates an executive-ready Excel workbook formatted to corporate manufacturing standards."""
    print(f"[REPORTING] Creating stylized Excel file at {output_path}...")
    wb = Workbook()
    
    # --- Sheet 1: OEE Dashboard ---
    ws1 = wb.active
    ws1.title = "OEE Metrics"
    ws1.views.sheetView[0].showGridLines = True
    
    # Theme Colors
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Corporate Blue
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_title = Font(name="Calibri", size=16, bold=True, color="1F497D")
    
    ws1.cell(row=2, column=2, value="Executive OEE Performance Log").font = font_title
    
    # Load DataFrame into sheet starting at Row 4
    for r in dataframe_to_rows(oee_df, index=False, header=True):
        ws1.append([] if ws1.max_row == 2 else r) # Gives spacing
        
    # Format Headers and Cells
    for col in range(1, oee_df.shape[1] + 1):
        cell = ws1.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center")
        
    # --- Sheet 2: Six Sigma Analytics ---
    ws2 = wb.create_sheet(title="Six Sigma SPC")
    ws2.views.sheetView[0].showGridLines = True
    ws2.cell(row=2, column=2, value="Statistical Process Control (SPC) Variables").font = font_title
    
    spc_rows = []
    for parameter, stats in spc_metrics.items():
        row_data = {"Parameter": parameter, **stats}
        spc_rows.append(row_data)
    df_spc = pd.DataFrame(spc_rows)
    
    for r in dataframe_to_rows(df_spc, index=False, header=True):
        ws2.append([] if ws2.max_row == 2 else r)
        
    for col in range(1, df_spc.shape[1] + 1):
        cell = ws2.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center")
        
    wb.save(output_path)
    print("[REPORTING] Excel sheet workbook saved successfully.")